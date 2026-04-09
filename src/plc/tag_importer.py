"""Bulk import parser for PLC point mapping files."""

from __future__ import annotations

import csv
import re
from collections import Counter
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


SUPPORTED_SUFFIXES = {".csv": "csv", ".xlsx": "excel"}
VALID_DATA_TYPES = {"float", "int", "bool", "string"}
IMPORTABLE_FIELDS = (
    "name",
    "address",
    "data_type",
    "unit",
    "description",
    "asset_id",
    "point_key",
    "deadband",
    "debounce_ms",
)
REQUIRED_FIELDS = ("name", "address")
TEMPLATE_HEADERS = list(IMPORTABLE_FIELDS)
TEMPLATE_SAMPLE_ROWS = [
    {
        "name": "pressure_diff",
        "address": "DB1.DBD0",
        "data_type": "float",
        "unit": "kPa",
        "description": "Main baghouse differential pressure",
        "asset_id": "ASSET_DUST_COLLECTOR_01",
        "point_key": "pressure_diff_kpa",
        "deadband": 0.2,
        "debounce_ms": 1000,
    },
    {
        "name": "dust_concentration",
        "address": "DB1.DBD4",
        "data_type": "float",
        "unit": "mg/m3",
        "description": "Outlet dust concentration",
        "asset_id": "ASSET_DUST_COLLECTOR_01",
        "point_key": "dust_concentration_mg_m3",
        "deadband": 0.1,
        "debounce_ms": 3000,
    },
]
TYPE_MISMATCH_PATTERNS = (
    ("bool", ("dbx", "x"), "Address suggests a bit/boolean point"),
    ("float", ("dbd",), "Address suggests a 32-bit numeric point"),
    ("int", ("dbw",), "Address suggests a 16-bit numeric point"),
)
POINT_SUGGESTION_CATALOG = {
    "pressure_diff_kpa": {
        "aliases": ("pressure_diff", "diff_pressure", "dp", "pressure"),
        "units": ("kpa",),
        "asset_id": "ASSET_DUST_COLLECTOR_01",
    },
    "fan_current_a": {
        "aliases": ("fan_current", "motor_current", "blower_current", "current"),
        "units": ("a", "amp", "amps"),
        "asset_id": "ASSET_DUST_COLLECTOR_01",
    },
    "airflow_m3h": {
        "aliases": ("airflow", "air_flow", "flowrate", "air_volume", "flow"),
        "units": ("m3/h", "m3h", "m3hr"),
        "asset_id": "ASSET_DUST_COLLECTOR_01",
    },
    "dust_concentration_mg_m3": {
        "aliases": ("dust_concentration", "dust", "outlet_dust", "emission"),
        "units": ("mg/m3", "mgm3"),
        "asset_id": "ASSET_DUST_COLLECTOR_01",
    },
    "cleaning_frequency_hz": {
        "aliases": ("cleaning_frequency", "cleaning", "pulse", "pulse_frequency"),
        "units": ("hz",),
        "asset_id": "ASSET_DUST_COLLECTOR_01",
    },
    "valve_state": {
        "aliases": ("valve_state", "valve", "damper", "state"),
        "units": (),
        "asset_id": "ASSET_DUST_COLLECTOR_01",
    },
    "temperature_c": {
        "aliases": ("temperature", "temp", "temp_c"),
        "units": ("c", "degc"),
        "asset_id": "ASSET_DUST_COLLECTOR_01",
    },
    "running_state": {
        "aliases": ("running_state", "running", "run_state", "run_status", "status"),
        "units": (),
        "asset_id": "ASSET_DUST_COLLECTOR_01",
    },
}
ASSET_SUGGESTION_CATALOG = {
    "ASSET_DUST_COLLECTOR_01": {
        "label": "1# Dust Collector",
        "scene_type": "dust",
        "point_keys": tuple(POINT_SUGGESTION_CATALOG.keys()),
    }
}

HEADER_ALIASES = {
    "name": {
        "name",
        "tag",
        "tagname",
        "tag_name",
        "point",
        "pointname",
        "point_name",
        "pointdisplayname",
        "variable",
        "variablename",
        "变量名",
        "点位",
        "点位名",
        "点位名称",
        "标签名",
        "标签名称",
    },
    "address": {
        "address",
        "addr",
        "plcaddress",
        "plc_address",
        "tagaddress",
        "tag_address",
        "register",
        "registeraddress",
        "寄存器",
        "地址",
        "点位地址",
        "plc地址",
    },
    "data_type": {
        "datatype",
        "data_type",
        "type",
        "valuetype",
        "value_type",
        "类型",
        "数据类型",
        "值类型",
    },
    "unit": {
        "unit",
        "units",
        "engineeringunit",
        "engineering_unit",
        "单位",
    },
    "description": {
        "description",
        "desc",
        "remark",
        "remarks",
        "comment",
        "memo",
        "metadata",
        "说明",
        "描述",
        "备注",
        "兼容元数据",
    },
    "asset_id": {
        "asset",
        "assetid",
        "asset_id",
        "assetcode",
        "asset_code",
        "equipmentasset",
        "资产",
        "资产id",
        "资产编号",
    },
    "point_key": {
        "key",
        "pointkey",
        "point_key",
        "semantickey",
        "semantic_key",
        "tagkey",
        "tag_key",
        "语义key",
        "点位key",
        "点位键",
        "监测点key",
    },
    "deadband": {
        "deadband",
        "delta",
        "delta_threshold",
        "threshold",
        "change_threshold",
        "死区",
        "变化阈值",
    },
    "debounce_ms": {
        "debounce",
        "debouncems",
        "debounce_ms",
        "anti_jitter_ms",
        "stabilityms",
        "防抖",
        "防抖ms",
        "防抖毫秒",
        "去抖毫秒",
    },
}

DATA_TYPE_ALIASES = {
    "float": "float",
    "double": "float",
    "real": "float",
    "number": "float",
    "decimal": "float",
    "int": "int",
    "integer": "int",
    "long": "int",
    "short": "int",
    "bool": "bool",
    "boolean": "bool",
    "bit": "bool",
    "string": "string",
    "str": "string",
    "text": "string",
}


@dataclass(slots=True)
class ParsedTabularData:
    columns: list[str]
    rows: list[tuple[int, dict[str, object]]]


@dataclass(slots=True)
class ParsedPreviewRow:
    row_number: int
    tag: dict[str, object]


def parse_device_tag_mapping_file(filename: str | None, content: bytes) -> dict[str, object]:
    """Parse a CSV or Excel file and normalize it into device tag rows."""
    return parse_device_tag_mapping_content(filename, content)


def parse_device_tag_mapping_content(
    filename: str | None,
    content: bytes,
    *,
    field_mapping: dict[str, str] | None = None,
    value_overrides: dict[int, dict[str, object]] | None = None,
) -> dict[str, object]:
    """Parse a CSV or Excel file with optional target-field to source-column mapping."""

    suffix = Path(filename or "").suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError("Only .csv and .xlsx files are supported for point mapping import")

    if suffix == ".csv":
        parsed = _read_csv(content)
    else:
        parsed = _read_excel(content)

    header_mapping, target_mapping = _build_header_mapping(parsed.columns, field_mapping=field_mapping)
    warnings: list[str] = []
    preview_rows: list[ParsedPreviewRow] = []
    skipped_rows = 0

    missing_fields = set(REQUIRED_FIELDS).difference(target_mapping.keys())
    if missing_fields:
        missing_display = ", ".join(sorted(missing_fields))
        raise ValueError(f"Missing required column(s): {missing_display}")

    for row_number, row in parsed.rows:
        normalized = _normalize_row(row_number, row, header_mapping, warnings)
        if normalized is None:
            skipped_rows += 1
            continue
        normalized = _apply_value_overrides(
            normalized,
            row_number=row_number,
            value_overrides=value_overrides,
            warnings=warnings,
        )
        preview_rows.append(ParsedPreviewRow(row_number=row_number, tag=normalized))

    validation_report = _build_validation_report(preview_rows)
    tags = [row.tag for row in preview_rows]
    serialized_preview_rows = [
        {
            "row_number": row.row_number,
            "tag": row.tag,
            **validation_report["rows_by_number"].get(row.row_number, {}),
        }
        for row in preview_rows
    ]

    return {
        "file_name": filename or "",
        "file_type": SUPPORTED_SUFFIXES[suffix],
        "detected_columns": parsed.columns,
        "matched_columns": header_mapping,
        "field_mapping": target_mapping,
        "unmatched_columns": [header for header in parsed.columns if header not in header_mapping],
        "available_fields": list(IMPORTABLE_FIELDS),
        "required_fields": list(REQUIRED_FIELDS),
        "total_rows": len(parsed.rows),
        "parsed_rows": len(tags),
        "skipped_rows": skipped_rows,
        "warnings": warnings,
        "tags": tags,
        "preview_rows": serialized_preview_rows,
        "validation_report": validation_report["summary"],
    }


def build_device_tag_import_template(file_type: str = "xlsx") -> tuple[bytes, str, str]:
    """Build a standard tag import template file."""

    normalized_type = file_type.lower()
    if normalized_type == "csv":
        return _build_csv_template()
    if normalized_type == "xlsx":
        return _build_excel_template()
    raise ValueError("Template format must be csv or xlsx")


def _read_csv(content: bytes) -> ParsedTabularData:
    decoded = None
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    if decoded is None:
        raise ValueError("Unable to decode CSV file") from last_error

    reader = csv.reader(StringIO(decoded))
    return _tabular_from_rows(reader)


def _read_excel(content: bytes) -> ParsedTabularData:
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    return _tabular_from_rows(rows)


def _tabular_from_rows(rows: Iterable[Iterable[object]]) -> ParsedTabularData:
    header_row: list[str] | None = None
    normalized_rows: list[tuple[int, dict[str, object]]] = []

    for index, raw_row in enumerate(rows, start=1):
        cells = list(raw_row)
        if header_row is None:
            if _is_empty_row(cells):
                continue
            header_row = [_to_header_string(cell) for cell in cells]
            continue

        if _is_empty_row(cells):
            continue

        row_map: dict[str, object] = {}
        for column_index, header in enumerate(header_row):
            if not header:
                continue
            row_map[header] = cells[column_index] if column_index < len(cells) else None
        normalized_rows.append((index, row_map))

    if header_row is None:
        raise ValueError("Import file is empty")

    return ParsedTabularData(columns=[header for header in header_row if header], rows=normalized_rows)


def _build_header_mapping(
    headers: Iterable[str],
    *,
    field_mapping: dict[str, str] | None = None,
) -> tuple[dict[str, str], dict[str, str]]:
    mapping: dict[str, str] = {}
    for header in headers:
        normalized_header = _normalize_header_key(header)
        for field_name, aliases in HEADER_ALIASES.items():
            if normalized_header in aliases:
                mapping[header] = field_name
                break

    target_mapping = _invert_mapping(mapping)
    if field_mapping:
        available_headers = {header for header in headers}
        next_mapping = {key: value for key, value in mapping.items() if value not in field_mapping}

        for field_name, source_column in field_mapping.items():
            if field_name not in IMPORTABLE_FIELDS:
                raise ValueError(f"Unsupported target field: {field_name}")
            if not source_column:
                target_mapping.pop(field_name, None)
                continue
            if source_column not in available_headers:
                raise ValueError(f"Unknown source column for mapping: {source_column}")
            next_mapping = {key: value for key, value in next_mapping.items() if value != field_name}
            next_mapping[source_column] = field_name

        mapping = next_mapping
        target_mapping = _invert_mapping(mapping)

    return mapping, target_mapping


def _normalize_row(
    row_number: int,
    row: dict[str, object],
    header_mapping: dict[str, str],
    warnings: list[str],
) -> dict[str, object] | None:
    values: dict[str, object] = {}
    for original_header, field_name in header_mapping.items():
        values[field_name] = row.get(original_header)

    name = _clean_string(values.get("name"))
    address = _clean_string(values.get("address"))
    if not name and not address:
        return None
    if not name or not address:
        warnings.append(f"Row {row_number}: missing required name or address, skipped")
        return None

    data_type = _normalize_data_type(values.get("data_type"))
    deadband = _parse_float(values.get("deadband"), row_number, "deadband", warnings)
    debounce_ms = _parse_int(values.get("debounce_ms"), row_number, "debounce_ms", warnings, default=0)

    return {
        "name": name,
        "address": address,
        "data_type": data_type,
        "unit": _clean_optional_string(values.get("unit")),
        "description": _clean_optional_string(values.get("description")),
        "asset_id": _clean_optional_string(values.get("asset_id")),
        "point_key": _clean_optional_string(values.get("point_key")),
        "deadband": deadband,
        "debounce_ms": debounce_ms,
    }


def _apply_value_overrides(
    tag: dict[str, object],
    *,
    row_number: int,
    value_overrides: dict[int, dict[str, object]] | None,
    warnings: list[str],
) -> dict[str, object]:
    overrides = (value_overrides or {}).get(row_number)
    if not overrides:
        return tag

    result = dict(tag)
    for field_name, raw_value in overrides.items():
        if field_name not in IMPORTABLE_FIELDS:
            warnings.append(f"Row {row_number}: override field '{field_name}' is not supported and was ignored")
            continue

        if field_name == "deadband":
            result[field_name] = _parse_float(raw_value, row_number, field_name, warnings)
            continue
        if field_name == "debounce_ms":
            result[field_name] = _parse_int(raw_value, row_number, field_name, warnings, default=0)
            continue
        if field_name == "data_type":
            result[field_name] = _normalize_data_type(raw_value)
            continue

        result[field_name] = _clean_optional_string(raw_value)

    return result


def _build_validation_report(preview_rows: list[ParsedPreviewRow]) -> dict[str, object]:
    address_counts: dict[str, int] = {}
    for row in preview_rows:
        address_key = _normalize_address_key(row.tag.get("address"))
        if not address_key:
            continue
        address_counts[address_key] = address_counts.get(address_key, 0) + 1

    suggestion_context = _build_suggestion_context(preview_rows)
    duplicate_clusters = _build_duplicate_address_clusters(preview_rows, address_counts)

    report_rows: list[dict[str, object]] = []
    issue_counts: dict[str, int] = {}
    error_count = 0
    warning_count = 0
    suggestion_count = 0

    for row in preview_rows:
        issues = _collect_row_validation_issues(row, address_counts)
        suggestions = _build_row_repair_suggestions(
            row,
            issues=issues,
            suggestion_context=suggestion_context,
        )
        flagged_fields = sorted({issue["field"] for issue in issues if issue.get("field")})
        status = "ok"
        if any(issue["severity"] == "error" for issue in issues):
            status = "error"
        elif issues:
            status = "warning"

        for issue in issues:
            issue_counts[issue["code"]] = issue_counts.get(issue["code"], 0) + 1
            if issue["severity"] == "error":
                error_count += 1
            else:
                warning_count += 1
        suggestion_count += len(suggestions)

        report_rows.append(
            {
                "row_number": row.row_number,
                "status": status,
                "flagged_fields": flagged_fields,
                "issues": issues,
                "suggestions": suggestions,
            }
        )

    rows_by_number = {
        row["row_number"]: {
            "status": row["status"],
            "flagged_fields": row["flagged_fields"],
            "issues": row["issues"],
            "suggestions": row["suggestions"],
        }
        for row in report_rows
    }

    return {
        "summary": {
            "total_rows": len(preview_rows),
            "clean_rows": sum(1 for row in report_rows if row["status"] == "ok"),
            "rows_with_errors": sum(1 for row in report_rows if row["status"] == "error"),
            "rows_with_warnings": sum(1 for row in report_rows if row["status"] == "warning"),
            "error_count": error_count,
            "warning_count": warning_count,
            "issue_counts": issue_counts,
            "suggestion_count": suggestion_count,
            "duplicate_clusters": duplicate_clusters,
            "has_errors": error_count > 0,
        },
        "rows_by_number": rows_by_number,
    }


def _collect_row_validation_issues(
    row: ParsedPreviewRow,
    address_counts: dict[str, int],
) -> list[dict[str, str]]:
    tag = row.tag
    issues: list[dict[str, str]] = []

    address_key = _normalize_address_key(tag.get("address"))
    if address_key and address_counts.get(address_key, 0) > 1:
        issues.append(
            _build_issue(
                code="duplicate_address",
                field="address",
                message="This PLC address is duplicated in the current import file.",
            )
        )

    if not _clean_string(tag.get("asset_id")):
        issues.append(
            _build_issue(
                code="missing_asset_id",
                field="asset_id",
                message="asset_id is empty. The point cannot be linked to an industrial asset.",
            )
        )

    if not _clean_string(tag.get("point_key")):
        issues.append(
            _build_issue(
                code="missing_point_key",
                field="point_key",
                message="point_key is empty. The point will not feed semantic diagnosis cleanly.",
            )
        )

    data_type = _clean_string(tag.get("data_type")).lower()
    address = _clean_string(tag.get("address"))
    if data_type not in VALID_DATA_TYPES:
        issues.append(
            _build_issue(
                code="unsupported_data_type",
                field="data_type",
                message=f"data_type '{data_type}' is not one of: {', '.join(sorted(VALID_DATA_TYPES))}.",
            )
        )
    else:
        mismatch = _detect_suspicious_type_mismatch(address, data_type)
        if mismatch:
            issues.append(
                _build_issue(
                    code="suspicious_type_mismatch",
                    field="data_type",
                    message=mismatch,
                )
            )

    return issues


def _build_suggestion_context(preview_rows: list[ParsedPreviewRow]) -> dict[str, object]:
    explicit_asset_counts = Counter(
        _clean_string(row.tag.get("asset_id"))
        for row in preview_rows
        if _clean_string(row.tag.get("asset_id"))
    )
    dominant_explicit_asset = explicit_asset_counts.most_common(1)[0][0] if explicit_asset_counts else None

    inferred_asset_counts = Counter()
    for row in preview_rows:
        point_match = _match_point_key_candidate(row.tag)
        if point_match:
            inferred_asset_counts[point_match["asset_id"]] += point_match["score"]

    dominant_inferred_asset = inferred_asset_counts.most_common(1)[0][0] if inferred_asset_counts else None

    return {
        "dominant_explicit_asset": dominant_explicit_asset,
        "dominant_inferred_asset": dominant_inferred_asset,
    }


def _build_row_repair_suggestions(
    row: ParsedPreviewRow,
    *,
    issues: list[dict[str, str]],
    suggestion_context: dict[str, object],
) -> list[dict[str, object]]:
    tag = row.tag
    suggestions: list[dict[str, object]] = []

    issue_codes = {issue["code"] for issue in issues}
    point_match = _match_point_key_candidate(tag)
    if "missing_point_key" in issue_codes and point_match:
        suggestions.append(
            {
                "field": "point_key",
                "value": point_match["point_key"],
                "confidence": point_match["confidence"],
                "reason": point_match["reason"],
            }
        )

    if "missing_asset_id" in issue_codes:
        asset_suggestion = _suggest_asset_id(
            tag,
            suggestion_context=suggestion_context,
            point_match=point_match,
        )
        if asset_suggestion:
            suggestions.append(asset_suggestion)

    if "duplicate_address" in issue_codes:
        cluster_key, cluster_label = _extract_address_cluster(tag.get("address"))
        suggestions.append(
            {
                "field": "address",
                "value": cluster_key,
                "confidence": "medium",
                "reason": f"Duplicate addresses are concentrated in {cluster_label}. Check whether this point should move to a nearby register in the same segment.",
            }
        )

    return suggestions


def _suggest_asset_id(
    tag: dict[str, object],
    *,
    suggestion_context: dict[str, object],
    point_match: dict[str, object] | None,
) -> dict[str, object] | None:
    explicit_asset = suggestion_context.get("dominant_explicit_asset")
    if explicit_asset:
        return {
            "field": "asset_id",
            "value": explicit_asset,
            "confidence": "high",
            "reason": f"Most mapped rows in this file already use {explicit_asset}.",
        }

    if point_match:
        return {
            "field": "asset_id",
            "value": point_match["asset_id"],
            "confidence": point_match["confidence"],
            "reason": f"{point_match['point_key']} usually belongs to {point_match['asset_id']}.",
        }

    inferred_asset = suggestion_context.get("dominant_inferred_asset")
    if inferred_asset:
        return {
            "field": "asset_id",
            "value": inferred_asset,
            "confidence": "medium",
            "reason": f"Nearby imported points look closest to {inferred_asset}.",
        }
    return None


def _match_point_key_candidate(tag: dict[str, object]) -> dict[str, object] | None:
    text_blob = " ".join(
        filter(
            None,
            [
                _normalize_lookup_text(tag.get("name")),
                _normalize_lookup_text(tag.get("description")),
                _normalize_lookup_text(tag.get("address")),
            ],
        )
    )
    normalized_unit = _normalize_unit(tag.get("unit"))

    best_match: dict[str, object] | None = None
    for point_key, metadata in POINT_SUGGESTION_CATALOG.items():
        score = 0
        for alias in metadata["aliases"]:
            normalized_alias = _normalize_lookup_text(alias)
            if not normalized_alias:
                continue
            if normalized_alias == _normalize_lookup_text(tag.get("name")):
                score += 5
            elif normalized_alias in text_blob:
                score += 3
        if normalized_unit and normalized_unit in { _normalize_unit(unit) for unit in metadata["units"] }:
            score += 2
        if score <= 0:
            continue

        confidence = "medium" if score < 6 else "high"
        candidate = {
            "point_key": point_key,
            "asset_id": metadata["asset_id"],
            "score": score,
            "confidence": confidence,
            "reason": f"The imported name/unit pattern matches {point_key}.",
        }
        if best_match is None or score > best_match["score"]:
            best_match = candidate

    return best_match


def _build_duplicate_address_clusters(
    preview_rows: list[ParsedPreviewRow],
    address_counts: dict[str, int],
) -> list[dict[str, object]]:
    clusters: dict[str, dict[str, object]] = {}
    for row in preview_rows:
        address = _clean_string(row.tag.get("address"))
        address_key = _normalize_address_key(address)
        if not address_key or address_counts.get(address_key, 0) <= 1:
            continue
        cluster_key, cluster_label = _extract_address_cluster(address)
        cluster = clusters.setdefault(
            cluster_key,
            {
                "cluster_key": cluster_key,
                "label": cluster_label,
                "addresses": set(),
                "row_numbers": set(),
            },
        )
        cluster["addresses"].add(address)
        cluster["row_numbers"].add(row.row_number)

    result = []
    for cluster in clusters.values():
        addresses = sorted(cluster["addresses"])
        row_numbers = sorted(cluster["row_numbers"])
        result.append(
            {
                "cluster_key": cluster["cluster_key"],
                "label": cluster["label"],
                "addresses": addresses,
                "row_numbers": row_numbers,
                "duplicate_count": len(addresses),
                "suggestion": f"Duplicate addresses are grouped in {cluster['label']}. Review rows {', '.join(str(value) for value in row_numbers)} together.",
            }
        )
    return sorted(result, key=lambda item: item["cluster_key"])


def _extract_address_cluster(address: object) -> tuple[str, str]:
    value = _clean_string(address)
    upper = value.upper()

    match = re.match(r"^(DB\d+)", upper)
    if match:
        cluster_key = match.group(1)
        return cluster_key, f"{cluster_key} register block"

    if upper.startswith("SIM:"):
        return "SIM", "SIM register segment"

    if upper.isdigit():
        number = int(upper)
        bucket_start = (number // 100) * 100
        bucket_end = bucket_start + 99
        cluster_key = f"{bucket_start:05d}-{bucket_end:05d}"
        return cluster_key, f"Modbus segment {cluster_key}"

    prefix = upper.split(".")[0].split(":")[0] if upper else "GENERAL"
    return prefix or "GENERAL", f"{prefix or 'GENERAL'} address segment"


def _build_issue(*, code: str, field: str, message: str, severity: str = "error") -> dict[str, str]:
    return {
        "code": code,
        "field": field,
        "message": message,
        "severity": severity,
    }


def _normalize_address_key(value: object) -> str:
    return _clean_string(value).upper()


def _detect_suspicious_type_mismatch(address: str, data_type: str) -> str | None:
    normalized_address = address.lower()
    for expected_type, markers, explanation in TYPE_MISMATCH_PATTERNS:
        if any(marker in normalized_address for marker in markers) and data_type != expected_type:
            return f"{explanation}, but data_type is '{data_type}'."
    return None


def _normalize_data_type(raw_value: object) -> str:
    text = _clean_string(raw_value).lower()
    if not text:
        return "float"
    return DATA_TYPE_ALIASES.get(text, text)


def _parse_float(raw_value: object, row_number: int, field_name: str, warnings: list[str]) -> float | None:
    text = _clean_string(raw_value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        warnings.append(f"Row {row_number}: invalid {field_name} value '{text}', ignored")
        return None


def _parse_int(
    raw_value: object,
    row_number: int,
    field_name: str,
    warnings: list[str],
    *,
    default: int,
) -> int:
    text = _clean_string(raw_value)
    if not text:
        return default
    try:
        return max(0, int(float(text)))
    except ValueError:
        warnings.append(f"Row {row_number}: invalid {field_name} value '{text}', defaulted to {default}")
        return default


def _clean_optional_string(value: object) -> str | None:
    text = _clean_string(value)
    return text or None


def _clean_string(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_header_string(value: object) -> str:
    return _clean_string(value)


def _normalize_lookup_text(value: object) -> str:
    lowered = _clean_string(value).lower()
    return re.sub(r"[^a-z0-9]+", " ", lowered).strip()


def _normalize_unit(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", _clean_string(value).lower())


def _normalize_header_key(value: str) -> str:
    normalized = _clean_string(value).lower()
    for token in (" ", "_", "-", ".", "(", ")", "[", "]", "{", "}", "/", "\\", ":", "："):
        normalized = normalized.replace(token, "")
    return normalized


def _is_empty_row(values: Iterable[object]) -> bool:
    return all(not _clean_string(value) for value in values)


def _invert_mapping(mapping: dict[str, str]) -> dict[str, str]:
    return {target: source for source, target in mapping.items()}


def _build_csv_template() -> tuple[bytes, str, str]:
    stream = StringIO()
    writer = csv.DictWriter(stream, fieldnames=TEMPLATE_HEADERS)
    writer.writeheader()
    writer.writerows(TEMPLATE_SAMPLE_ROWS)
    content = stream.getvalue().encode("utf-8-sig")
    return content, "device_tag_import_template.csv", "text/csv; charset=utf-8"


def _build_excel_template() -> tuple[bytes, str, str]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PointMapping"
    sheet.append(TEMPLATE_HEADERS)
    for row in TEMPLATE_SAMPLE_ROWS:
        sheet.append([row.get(header, "") for header in TEMPLATE_HEADERS])
    sheet.freeze_panes = "A2"

    description_sheet = workbook.create_sheet("Instructions")
    description_sheet.append(["field", "required", "description"])
    descriptions = {
        "name": "PLC point display name or unique tag name",
        "address": "PLC address such as DB1.DBD0, 40001, SIM:1",
        "data_type": "float, int, bool, string",
        "unit": "Engineering unit shown in the UI",
        "description": "Optional metadata or notes",
        "asset_id": "Asset bound to intelligence patrol",
        "point_key": "Semantic point key used by diagnostics",
        "deadband": "Only keep history when change exceeds this threshold",
        "debounce_ms": "Minimum stable time before change is accepted",
    }
    for field_name in TEMPLATE_HEADERS:
        description_sheet.append(
            [
                field_name,
                "yes" if field_name in REQUIRED_FIELDS else "no",
                descriptions[field_name],
            ]
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return (
        buffer.getvalue(),
        "device_tag_import_template.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
